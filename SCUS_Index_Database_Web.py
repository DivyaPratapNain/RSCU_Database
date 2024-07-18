from flask import Flask, render_template, request, jsonify , send_file
from io import BytesIO
import openpyxl
import lxml  # Ensure lxml is installed
import os
import pandas as pd
import sqlite3

app = Flask(__name__)
app.secret_key = 'DOST'

# Define the directory where uploaded files will be saved
upload_dir = 'UPLOADS'
app.config['UPLOAD_FOLDER'] = upload_dir

# Read the Excel file and convert to SQLite database
excel_file_path = os.path.join(upload_dir, 'Plant_RSCU.xlsx')
df = pd.read_excel(excel_file_path)

# Convert DataFrame to SQLite
conn = sqlite3.connect(os.path.join(upload_dir, 'scus_index.db'))
df.to_sql('scus_index', conn, if_exists='replace', index=False)
conn.close()

@app.route('/')
def home():
    conn = sqlite3.connect(os.path.join(upload_dir, 'scus_index.db'))
    df = pd.read_sql_query("SELECT * FROM scus_index", conn)
    species = sorted(df['species'].unique().tolist())
    data = df.to_html(classes='table table-bordered table-centered', index=False)
    conn.close()
    return render_template('home.html', table_data=data, df_columns=df.columns, species=species)

@app.route('/filter', methods=['POST'])
def filter_data():
    filters = request.json['filters']
    selected_species = filters.get('species', [])

    if not selected_species:
        return jsonify(data="No Such Combination")

    conn = sqlite3.connect(os.path.join(upload_dir, 'scus_index.db'))
    query = "SELECT * FROM scus_index WHERE species IN ({})".format(','.join('?' * len(selected_species)))
    df = pd.read_sql_query(query, conn, params=selected_species)
    conn.close()

    if df.empty:
        return jsonify(data="No Such Combination")

    # Ensure the 'frequency' column is numeric
    df['frequency'] = pd.to_numeric(df['frequency'], errors='coerce')

    # Calculate the average RSCU values for each codon
    avg_df = df.groupby('codon', as_index=False)['frequency'].mean()
    avg_df.columns = ['Codon', 'Average RSCU']

    data = avg_df.to_html(classes='table table-bordered table-centered', index=False)
    return jsonify(data=data, selected_species=selected_species)

@app.route('/get_species', methods=['GET'])
def get_species():
    conn = sqlite3.connect(os.path.join(upload_dir, 'scus_index.db'))
    df = pd.read_sql_query("SELECT DISTINCT species FROM scus_index", conn)
    species = sorted(df['species'].tolist())
    conn.close()
    return jsonify(species=species)

@app.route('/download', methods=['POST'])
def download_file():
    data = request.json.get('data')
    filename = request.json.get('filename')
    
    # Check if the filename is 'RSCU_Database.xlsx' and send 'Plant_RSCU.xlsx' directly
    if filename == 'RSCU_Database.xlsx':
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Plant_RSCU.xlsx')
        if not os.path.exists(file_path):
            return jsonify({"error": "File not found"}), 404
        return send_file(file_path, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    # Ensure data is a valid HTML table string for 'Average_RSCU.xlsx'
    if data and filename == 'Average_RSCU.xlsx':
        try:
            # Convert HTML table to DataFrame
            df = pd.read_html(data)[0]
        except ValueError as e:
            return jsonify({"error": f"Error parsing HTML: {e}"}), 400
        except IndexError as e:
            return jsonify({"error": f"Error processing HTML: {e}"}), 400

        # Load the existing Excel file
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        if not os.path.exists(file_path):
            return jsonify({"error": "File not found"}), 404

        # Load the workbook and select the active worksheet
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Clear the existing content
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.value = None

        # Write new DataFrame to the worksheet
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Save the workbook to a BytesIO object
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        return jsonify({"error": "No data or filename provided"}), 400
    
if __name__ == '__main__':
    app.run(debug=True)
